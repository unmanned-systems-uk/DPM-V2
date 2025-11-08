# coding:utf-8
import sys
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import re

lst_model_pair = { \
    "HC":"ILCE-1M2",
    "TK":"ILCE-1",
    "OS/GS":"ILCE-9M3",
    "TP":"ILCE-9M2",
    "GL":"ILCE-7RM5",
    "MGA":"ILCE-7RM4A",
    "MG":"ILCE-7RM4",
    "SP/GLS":"ILCE-7CR",
    "OT":"ILCE-7SM3",
    "LS":"ILCE-7M4",
    "PP/NB2":"ILCE-7CM2",
    "NB":"ILCE-7C",
    "DZ":"ILCE-6700",
    "LOSGATOS":"MPC-2610",
    "Chaplin":"ILME-FX6",
    "ALCINA":"ILME-FX3A",
    "ALCIN":"ILME-FX3",
    "LAX":"ILME-FX30",
    "Harry":"PXW-Z200",
    "Potter":"HXR-NX800",
    "SA":"ZV-E1",
    "NE":"ZV-E10M2",
    "WJ":"DSC-RX0M2",
}

lst_model_seq = [
    "ILCE-1M2",
    "ILCE-1",
    "ILCE-9M3",
    "ILCE-9M2",
    "ILCE-7RM5",
    "ILCE-7RM4A",
    "ILCE-7RM4",
    "ILCE-7CR",
    "ILCE-7SM3",
    "ILCE-7M4",
    "ILCE-7CM2",
    "ILCE-7C",
    "ILCE-6700",
    "MPC-2610",
    "ILME-FX6",
    "ILME-FX3A",
    "ILME-FX3",
    "ILME-FX30",
    "PXW-Z200",
    "HXR-NX800",
    "ZV-E1",
    "ZV-E10M2",
    "DSC-RX0M2",
]

def main(excel_path, ver_make):
    # 1. コマンド一覧を開く
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["CrSDK コマンド対応表"]

    file_output = open("functionList.csv",mode="w",encoding="utf-8")
    file_output.write("No.\tAPIs\tMode\tILCE-1M2\tILCE-1\tILCE-9M3\tILCE-9M2\tILCE-7RM5\tILCE-7RM4A\tILCE-7RM4\tILCE-7CR\tILCE-7SM3\tILCE-7M4\tILCE-7CM2\tILCE-7C\tILCE-6700\tMPC-2610\tILME-FX6\tILME-FX3A\tILME-FX3\tILME-FX30\tPXW-Z200\tHXR-NX800\tZV-E1\tZV-E10M2\tDSC-RX0M2\n")
    lst_not_process_PTP = []
    num_process_feature = 0

    # 2. データを走査する

    # バージョン名と列を記憶しておく
    lst_version_and_column_letter = {}
    column_index_start_version = column_index_from_string("V")
    for cell_ver in ws["9"]:
        if cell_ver.column < column_index_start_version:
            continue
        if cell_ver.value is None:
            break
        version_codename = cell_ver.value
        # 空白と改行を除去
        version_codename = version_codename.strip().replace('\n','')
        # 先頭のアルファベットのみを使う
        match = re.search('[a-zA-Z]+',version_codename)
        version_codename = match.group(0)

        cell_ver_num = ws[f"{cell_ver.column_letter}{10}"]

        lst_version_and_column_letter[cell_ver_num.value] = {"ver_codename":version_codename, "column_letter":cell_ver.column_letter}

    # モデル名と列を記憶しておく
    lst_model_and_column_letter = {}
    column_index_start_model = column_index_from_string("AH")
    for cell_model in ws["9"]:
        if cell_model.column < column_index_start_model:
            continue
        if cell_model.value is None:
            break

        model = cell_model.value
        model = model.split('\n')[0]
        lst_model_and_column_letter[model] = cell_model.column_letter
    # print(f"lst_model_and_column_letter:{lst_model_and_column_letter}")

    min_row = ws.min_row
    max_row = ws.max_row
    for r in range(min_row,max_row+1):
        # B13まではスキップする
        if r < 13:
            continue
        # F列が空じゃないものはスキップする
        cell = ws[f"F{r}"]
        if cell.value is not None:
            continue
        
        # 作成したいバージョンでは「○」になっていない機能はスキップする
        column_letter_by_ver = lst_version_and_column_letter[ver_make]["column_letter"]
        # print(f"column_letter_by_ver:{column_letter_by_ver}")
        cell_feature_by_ver = ws[f"{column_letter_by_ver}{r}"]
        # print(f"cell_feature_by_ver:{cell_feature_by_ver.value}")
        if cell_feature_by_ver.value != '○':
            cell_feature_name = ws[f"C{r}"]
            cell_ptp_code = ws[f"D{r}"]
            if cell_feature_name.value is not None:
                pass
            continue

        # Codeと一致する機能の公開名を対応表(別ファイル)から取得
        # 生成するCSVファイルに必要な情報を書く
        cell_feature_name = ws[f"C{r}"]
        cell_ptp_code = ws[f"D{r}"]
        # print(f"feature_name:{cell_feature_name.value}, code={cell_ptp_code.value}")

        # モデル毎の対応データを出力する
        data_models = ""
        for model_seq in lst_model_seq:
            column_letter = lst_model_and_column_letter[model_seq]
            cell_valid = ws[f"{column_letter}{r}"]
            value_valid = cell_valid.value
            if value_valid == "○":
                value_valid = "○"
            elif value_valid == "×":
                value_valid = "×"

            data_models += value_valid + "\t"
        
        # TODO: modeを算出する
        mode = ""

        name_feature = cell_feature_name.value
        name_feature = name_feature.replace('\n','')
        name_feature = name_feature.replace('\t',' ')

        file_output.write(f"{num_process_feature+1}\t{name_feature}\t{mode}\t{data_models}\n")
        num_process_feature += 1

    # 3. 処理しなかった機能についてログ出力する
    if len(lst_not_process_PTP) != 0:
        # TODO: 表示処理を実装
        pass
    pass

if __name__ == "__main__":
    # 引数に、以下を要求する
    # 1.元Excelファイルのパス
    # 2.作成するドキュメントのバージョン
    if len(sys.argv)<3:
        print("usage generate_functionList.py [excel_path] [ver]")
        exit(-1)
    excel_path = sys.argv[1]
    ver = sys.argv[2]

    main(excel_path, ver)
